"""US life table → per-year micromorts for select ages.

Source: CDC NCHS NVSR 74(2), 'United States Life Tables, 2022', Table 1
(total population). The XLSX is downloaded directly; we keep an
embedded fallback of the qx values for ages we publish so the scraper
works offline.
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO

import requests

from micromort.convert import from_probability
from micromort.db import add_tags, transaction, upsert_risk, upsert_source
from scrape._http import USER_AGENT, fetch  # fetch unused but warms cache

SOURCE = dict(
    name="CDC NCHS — US Life Tables, 2022 (NVSR 74-02 Table 1)",
    url="https://ftp.cdc.gov/pub/Health_Statistics/NCHS/Publications/NVSR/74-02/Table01.xlsx",
    publisher="US CDC / NCHS",
)

# Ages we publish. We pick a representative sweep so the chart shows the
# classic mortality curve without 100 redundant rows.
AGES = [0, 1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95]

# Fallback qx values (probability of dying within one year) from the same
# table, used if the XLSX can't be fetched. Two-sig-fig fidelity.
FALLBACK_QX: dict[int, float] = {
    0:  0.005605,
    1:  0.000444,
    5:  0.000155,
    10: 0.000101,
    15: 0.000402,
    20: 0.001188,
    25: 0.001724,
    30: 0.002034,
    35: 0.002463,
    40: 0.002862,
    45: 0.003967,
    50: 0.005776,
    55: 0.008601,
    60: 0.012760,
    65: 0.018850,
    70: 0.029137,
    75: 0.045625,
    80: 0.075028,
    85: 0.128037,
    90: 0.215404,
    95: 0.339875,
}


def _try_fetch_qx() -> dict[int, float] | None:
    try:
        resp = requests.get(SOURCE["url"], headers={"User-Agent": USER_AGENT}, timeout=30)
        resp.raise_for_status()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content), data_only=True)
        ws = wb.active
        out: dict[int, float] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            age_label, qx = row[0], row[1]
            if not age_label or qx is None:
                continue
            try:
                age = int(str(age_label).split("–")[0].split("-")[0].strip())
            except ValueError:
                continue
            if age in AGES:
                out[age] = float(qx)
        return out if len(out) == len(AGES) else None
    except Exception:
        return None


def ingest(conn) -> int:
    qx = _try_fetch_qx()
    used_live = qx is not None
    if qx is None:
        qx = FALLBACK_QX

    source_id = upsert_source(
        conn,
        accessed_at=dt.date.today().isoformat() if used_live else None,
        **SOURCE,
        notes=(SOURCE.get("notes") or "") + (
            " Loaded from live CDC FTP." if used_live else " Loaded from embedded fallback."
        ),
    )

    n = 0
    with transaction(conn):
        for age in AGES:
            mm = from_probability(qx[age])
            label = f"All-cause mortality at age {age}"
            risk_id = upsert_risk(
                conn,
                slug=f"cdc-life:age-{age:03d}-us-2022",
                name=f"{label} (US, 2022)",
                category="baseline today",
                micromorts=mm,
                exposure="per_year",
                exposure_detail=f"Probability of dying within one year, exact age {age}, 2022",
                population="US, both sexes",
                region="US",
                year=2022,
                source_id=source_id,
                original_value=f"qx = {qx[age]:.6f}",
                original_unit="qx (probability of dying within one year)",
                confidence="high",
            )
            add_tags(conn, risk_id, ["baseline", "age", f"age-{age}"])
            n += 1
    return n


if __name__ == "__main__":
    from micromort.db import connect, init_schema
    conn = connect()
    init_schema(conn)
    print(f"CDC life table: ingested {ingest(conn)} entries.")
